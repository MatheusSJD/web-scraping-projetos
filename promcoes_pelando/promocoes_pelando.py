from selenium import webdriver
import os
import openpyxl
import time


class BuscaPrecos:
    def __init__(self):
        self.driver = webdriver.Chrome(
            executable_path=os.getcwd() + os.sep + 'chromedriver.exe')

    def iniciar(self):
        self.numero_pagina_atual = 1
        self.driver.get(
            f'https://www.pelando.com.br/grupo/pc-gamer-quente')
        self.criar_planilha()
        self.encontrar_promocoes()

    def criar_planilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('Promocões PC')
        self.planilha_de_promocoes = self.planilha['Promocões PC']
        self.planilha_de_promocoes.cell(row=1, column=1, value="Títulos")
        self.planilha_de_promocoes.cell(row=1, column=2, value="Preços")
        self.planilha_de_promocoes.cell(row=1, column=3, value="Lojas")
        self.planilha_de_promocoes.cell(row=1, column=4, value="Datas")
        self.planilha_de_promocoes.cell(row=1, column=5, value="Links")

    def encontrar_promocoes(self):
        try:
            while True:
                # título
                self.titulos = self.driver.find_elements_by_xpath(
                    "//a[@class='cept-tt thread-link linkPlain thread-title--card']")
                # Preços
                self.precos = self.driver.find_elements_by_xpath(
                    "//span[@class='thread-price text--b cept-tp size--all-l']")
                # Lojas
                self.lojas = self.driver.find_elements_by_xpath(
                    "//span[@class='cept-merchant-name text--b text--color-brandPrimary link']")
                # Datas
                self.datas = self.driver.find_elements_by_xpath(
                    "//span[@class='hide--toBigCards1']")

                self.armazenar_promocoes()
                self.proxima_pagina()

        except Exception as erro:
            print('Não há mais pesquisas para extrair ddaqui não moço')
            print(erro)

    def proxima_pagina(self):
        self.numero_pagina_atual += 1
        self.driver.get(
            f'https://www.pelando.com.br/grupo/pc-gamer-quente?page={self.numero_pagina_atual}')
        if self.numero_pagina_atual == 3:
            print('Acabei de filtrar essa joça moço')
            self.driver.quit()

    def armazenar_promocoes(self):
        for indice in range(0, len(self.titulos)):
            nova_linha = [
                self.titulos[indice].text,
                self.precos[indice].text,
                self.lojas[indice].text,
                self.datas[indice].text,
                self.titulos[indice].get_attribute("href")
            ]
            self.planilha_de_promocoes.append(nova_linha)
            self.planilha.save('promocoes_pelando.xlsx')


precos = BuscaPrecos()
precos.iniciar()